#!/usr/bin/env python
"""
需传入 1 个参数：
    1. 项目路径（包含 headMessage*.xlsx、sampleInfo.xlsx、results、figures 等）

功能：
    - 读取 headMessage，识别项目类型（AA/FFA/OA 等）
    - 根据《小高靶合并报告文件需求》在项目路径下生成 exp-data/ 目录
    - 复制指定 QC / SPL 图及必要的结果表
    - 在项目路径下生成 exp-data.zip
"""

import os
import sys
import glob
import shutil
import zipfile

import pandas as pd
from openpyxl import load_workbook


# ---------- 工具函数 ----------

def detect_scheme(product_type: str) -> str | None:
    """根据 headMessage 中的 '项目类型' 映射到打包方案号。"""
    if product_type is None:
        return None

    pt = str(product_type).strip()

    # 小靶标方案一：AA / AA-Pro / ARA / TRP / BA / PH / NT
    scheme1 = {
        "氨基酸", "氨基酸Pro", "氧化脂质", "色氨酸",
        "胆汁酸", "植物激素", "神经递质",
        "AA", "AA-Pro", "ARA", "TRP", "BA", "PH", "NT",
    }

    # 小靶标方案二：FFA / SCFA
    scheme2 = {
        "游离脂肪酸", "短酸",
        "FFA", "SCFA",
    }

    # 小靶标方案三：OA / CCM
    scheme3 = {
        "有机酸", "能量代谢",
        "OA", "CCM",
    }

    # 大高靶方案一：AQ700、AQ-血液500、AQ-1000（RP+HILIC，带RT图）
    scheme_aq_rp_hilic = {
        "AQ700", "AQ-血液500", "AQ-1000", "AQ1000",
    }

    # 大高靶方案二：肠菌300（GCMS+LCMS 或仅 LCMS）
    scheme_gm = {
        "肠菌300", "GM300", "肠菌300-LC",
    }

    # 大高靶方案三：暴露组（单组，带TIC+RT+QC-corrplot）
    scheme_exposome = {
        "暴露组", "Exposome",
    }

    # 定量脂质（单组，TIC+IS，带QC-corrplot）
    scheme_lipids = {
        "定量脂质", "Lipids",
    }

    if pt in scheme1:
        return "scheme1"
    if pt in scheme2:
        return "scheme2"
    if pt in scheme3:
        return "scheme3"
    if pt in scheme_aq_rp_hilic:
        return "scheme_aq_rp_hilic"
    if pt in scheme_gm:
        return "scheme_gm"
    if pt in scheme_exposome:
        return "scheme_exposome"
    if pt in scheme_lipids:
        return "scheme_lipids"

    return None


def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


def get_first_sample_name(in_dir: str, sample_type: str) -> tuple[str | None, int]:
    """从 sampleInfo.xlsx 中读取指定 sample_type 的第一个 sample_name。

    Returns:
        tuple[str | None, int]: (样本名称, 样本数量)
    """
    sample_info_path = os.path.join(in_dir, "sampleInfo.xlsx")
    if not os.path.isfile(sample_info_path):
        print(f"ExpDataPack: sampleInfo.xlsx 不存在：{sample_info_path}")
        return None, 0

    try:
        df = pd.read_excel(sample_info_path, sheet_name=0, engine='openpyxl')
        filtered = df[df["sample_type"] == sample_type]
        sample_count = len(filtered)

        if sample_count > 0:
            sample_name = str(filtered.iloc[0]["sample_name"])
            print(
                f"ExpDataPack: 找到 {sample_count} 个 {sample_type} 样本，第一个为：{sample_name}")
            return sample_name, sample_count
        else:
            print(
                f"ExpDataPack: sampleInfo.xlsx 中没有 sample_type='{sample_type}' 的样本")
            return None, 0
    except Exception as e:
        print(f"ExpDataPack: 读取 sampleInfo.xlsx 失败：{e}")

    return None, 0


def find_figure_file(base_path: str, sample_name: str, extensions=(".pdf", ".png")) -> str | None:
    """在 base_path 下查找 sample_name.{ext} 的文件，优先级按 extensions 顺序。"""
    if not os.path.isdir(base_path):
        print(f"ExpDataPack: 目录不存在：{base_path}")
        return None

    for ext in extensions:
        candidate = os.path.join(base_path, f"{sample_name}{ext}")
        print(f"ExpDataPack: 尝试查找图片：{candidate}")
        if os.path.isfile(candidate):
            print(f"ExpDataPack: 找到图片：{candidate}")
            return candidate

    print(f"ExpDataPack: 未找到 {sample_name} 对应的图片文件（扩展名：{extensions}）")
    return None


def pack_qc_spl_for_dir(in_dir: str, dest_root: str):
    """
    通用：给某一个 in_dir（一个项目或一个 RP/HILIC 子项目）
    在 dest_root 下生成 QC/QC.png 和 SPL/SPL.png。

    参考 Rmd 模板逻辑：
      - QC 图：从 sampleInfo 读第一个 QC 的 sample_name，拼路径 figures/XIC/QC/{sample_name}.pdf/png
        若找不到则退化到 Calibrator/results/figures/XIC/QC/HQC.png
      - SPL 图：从 sampleInfo 读第一个 SPL 的 sample_name，拼路径 figures/TIC/SPL/{sample_name}.png
    """
    qc_dir = os.path.join(dest_root, "QC")
    spl_dir = os.path.join(dest_root, "SPL")
    ensure_dir(qc_dir)
    ensure_dir(spl_dir)

    # ===== QC 图 =====
    qc_sample_name, qc_count = get_first_sample_name(in_dir, "QC")
    qc_fig_path = None

    # QC 数量大于 1 才算有 QC，否则算无 QC 处理
    if qc_sample_name and qc_count > 1:
        # 优先在 results/figures/XIC/QC 下找 PNG
        src_qc_dir = os.path.join(in_dir, "results", "figures", "XIC", "QC")
        qc_fig_path = find_figure_file(
            src_qc_dir, qc_sample_name, extensions=(".png",))  # 只找 PNG

        # 如果 results/figures 下找不到，再试 figures/XIC/QC
        if not qc_fig_path:
            src_qc_dir = os.path.join(in_dir, "figures", "XIC", "QC")
            qc_fig_path = find_figure_file(
                src_qc_dir, qc_sample_name, extensions=(".png",))  # 只找 PNG
    else:
        if qc_count <= 1:
            print(f"ExpDataPack: QC 数量为 {qc_count}，不满足有 QC 的条件（需 > 1）")

    # 如果没有 QC 或找不到对应图，退化到 HQC.png
    # Calibrator 文件夹与 in_dir 同级
    if not qc_fig_path:
        parent_dir = os.path.dirname(in_dir)
        fallback_qc = os.path.join(
            parent_dir, "Calibrator", "results", "figures", "XIC", "QC", "HQC.png"
        )
        print(f"ExpDataPack: 尝试使用 HQC 备用图：{fallback_qc}")
        if os.path.isfile(fallback_qc):
            qc_fig_path = fallback_qc
            print(f"ExpDataPack: 使用 HQC 图：{fallback_qc}")
        else:
            print(f"ExpDataPack: HQC 备用图也不存在")

    # 复制到 dest_root/QC/QC.png
    if qc_fig_path:
        dest_qc = os.path.join(qc_dir, "QC.png")
        print(f"ExpDataPack: 复制 QC 图：{qc_fig_path} -> {dest_qc}")
        shutil.copy2(qc_fig_path, dest_qc)
    else:
        print(f"ExpDataPack: 未找到任何 QC 图，跳过")

    # ===== SPL 图 =====
    spl_sample_name, spl_count = get_first_sample_name(in_dir, "SPL")
    spl_fig_path = None

    if spl_sample_name:
        # 优先在 results/figures/TIC/SPL 下找 PNG
        src_spl_dir = os.path.join(in_dir, "results", "figures", "TIC", "SPL")
        spl_fig_path = find_figure_file(
            src_spl_dir, spl_sample_name, extensions=(".png",))  # 只找 PNG

        # 如果 results/figures 下找不到，再试 figures/TIC/SPL
        if not spl_fig_path:
            src_spl_dir = os.path.join(in_dir, "figures", "TIC", "SPL")
            spl_fig_path = find_figure_file(
                src_spl_dir, spl_sample_name, extensions=(".png",))  # 只找 PNG

    # 复制到 dest_root/SPL/SPL.png
    if spl_fig_path:
        dest_spl = os.path.join(spl_dir, "SPL.png")
        print(f"ExpDataPack: 复制 SPL 图：{spl_fig_path} -> {dest_spl}")
        shutil.copy2(spl_fig_path, dest_spl)
    else:
        print(f"ExpDataPack: 未找到任何 SPL 图，跳过")


# ---------- 各方案具体打包 ----------

def pack_scheme1(in_dir: str, exp_dir: str):
    """
    方案一：AA / AA-Pro / ARA / TRP / BA / PH / NT

    结构：
    exp-data/
      ├─ QC/QC.png
      └─ SPL/SPL.png
    """
    pack_qc_spl_for_dir(in_dir, exp_dir)


def pack_scheme2(in_dir: str, exp_dir: str):
    """
    方案二：FFA / SCFA

    结构：
    exp-data/
      ├─ QC/QC.png
      ├─ SPL/SPL.png
      ├─ Calibrators.best.xlsx
      └─ Quantification-Stats.xlsx
    """
    pack_qc_spl_for_dir(in_dir, exp_dir)

    results_dir = os.path.join(in_dir, "results")
    files_to_copy = [
        ("Calibrators.best.xlsx", "Calibrators.best.xlsx"),
        ("Quantification-Stats.xlsx", "Quantification-Stats.xlsx"),
    ]

    for src_name, dest_name in files_to_copy:
        src_path = os.path.join(results_dir, src_name)
        if os.path.isfile(src_path):
            shutil.copy2(src_path, os.path.join(exp_dir, dest_name))


def pack_scheme3(in_dir: str, exp_dir: str, grp_type: str | None):
    """
    小靶标方案三：OA / CCM（RP + HILIC）

    结构：
    exp-data/
      ├─ RP/QC/QC.png, RP/SPL/SPL.png
      └─ HILIC/QC/QC.png, HILIC/SPL/SPL.png
    """
    if not grp_type:
        grp_type = "RP|HILIC"

    grps = [g.strip() for g in str(grp_type).split("|") if g.strip()]
    if not grps:
        return

    for grp in grps:
        # 按 FeishuRecord.py 的思路，将路径中的 grp_type 替换为当前 grp
        try:
            import re
            grp_in_dir = re.sub(grp_type, grp, in_dir)
        except Exception:
            grp_in_dir = in_dir

        dest_root = os.path.join(exp_dir, grp)
        pack_qc_spl_for_dir(grp_in_dir, dest_root)


def pack_tic_with_modes(in_dir: str, tic_dir: str, sample_type: str, subdir_name: str):
    """
    通用：打包 TIC 图，支持 pos/neg 模式
    用于定量脂质和暴露组
    """
    dest_dir = os.path.join(tic_dir, subdir_name)
    ensure_dir(dest_dir)

    sample_name, sample_count = get_first_sample_name(in_dir, sample_type)
    if not sample_name:
        return

    # 查找 TIC 图：可能有 pos/neg 后缀
    src_tic_dir = os.path.join(
        in_dir, "results", "figures", "TIC", sample_type)
    if not os.path.isdir(src_tic_dir):
        src_tic_dir = os.path.join(in_dir, "figures", "TIC", sample_type)

    if os.path.isdir(src_tic_dir):
        # 查找所有 png 文件
        for mode_suffix in ["", "__pos", "__pos1", "__neg"]:
            fig_name = f"{sample_name}{mode_suffix}.png"
            src_path = os.path.join(src_tic_dir, fig_name)
            if os.path.isfile(src_path):
                # 生成目标文件名
                dest_name = f"{subdir_name}{mode_suffix}.png"
                dest_path = os.path.join(dest_dir, dest_name)
                print(f"ExpDataPack: 复制 TIC 图：{src_path} -> {dest_path}")
                shutil.copy2(src_path, dest_path)


def copy_rt_is_files(in_dir: str, dest_dir: str):
    """
    通用：复制 RT 和 IS 相关文件
    """
    results_dir = os.path.join(in_dir, "results")

    # RT 图
    rt_figs = [
        ("figures/IS/RT-Control-Positive__0.png", "RT-Control-Positive__0.png"),
        ("figures/IS/RT-Control-Negative__0.png", "RT-Control-Negative__0.png"),
    ]

    for rel_src, dest_name in rt_figs:
        src_path = os.path.join(results_dir, rel_src)
        if os.path.isfile(src_path):
            dest_path = os.path.join(dest_dir, dest_name)
            print(f"ExpDataPack: 复制 RT 图：{src_path} -> {dest_path}")
            shutil.copy2(src_path, dest_path)


def pack_scheme_aq_rp_hilic(in_dir: str, exp_dir: str, grp_type: str | None):
    """
    大高靶方案一：AQ700、AQ-血液500、AQ-1000

    结构：
    exp-data/
      ├─ HILIC/QC/QC.png, SPL/SPL.png, RT/RT-Control-*.png
      ├─ RP/QC/QC.png, SPL/SPL.png, RT/RT-Control-*.png
      ├─ QC-corrplot-HILIC.jpg
      ├─ QC-corrplot-RP.jpg
      └─ Quantification-IS.xlsx
    """
    if not grp_type:
        grp_type = "RP|HILIC"

    grps = [g.strip() for g in str(grp_type).split("|") if g.strip()]
    if not grps:
        return

    for grp in grps:
        try:
            import re
            grp_in_dir = re.sub(grp_type, grp, in_dir)
        except Exception:
            grp_in_dir = in_dir

        # QC / SPL 图
        grp_root = os.path.join(exp_dir, grp)
        pack_qc_spl_for_dir(grp_in_dir, grp_root)

        # RT 图
        rt_dir = os.path.join(grp_root, "RT")
        ensure_dir(rt_dir)
        copy_rt_is_files(grp_in_dir, rt_dir)

    # QC-corrplot 图（放在 exp_dir 根目录，与 Quantification-IS.xlsx 同级）
    for grp in grps:
        try:
            import re
            grp_in_dir = re.sub(grp_type, grp, in_dir)
        except Exception:
            grp_in_dir = in_dir

        results_dir = os.path.join(grp_in_dir, "results")
        # 源文件已经带有 grp 后缀，直接查找完整文件名
        src_filename = f"QC-corrplot-{grp}.jpg"
        qc_corrplot_src = os.path.join(results_dir, "figures", src_filename)
        if os.path.isfile(qc_corrplot_src):
            dest_path = os.path.join(exp_dir, src_filename)
            print(
                f"ExpDataPack: 复制 QC-corrplot：{qc_corrplot_src} -> {dest_path}")
            shutil.copy2(qc_corrplot_src, dest_path)
        else:
            print(f"ExpDataPack: 未找到 QC-corrplot 图：{qc_corrplot_src}")

    # Quantification-IS.xlsx（只需一份，从第一个 grp 取）
    first_grp = grps[0]
    try:
        import re
        first_grp_dir = re.sub(grp_type, first_grp, in_dir)
    except Exception:
        first_grp_dir = in_dir

    quant_is_src = os.path.join(
        first_grp_dir, "results", "Quantification-IS.xlsx")
    if os.path.isfile(quant_is_src):
        dest_path = os.path.join(exp_dir, "Quantification-IS.xlsx")
        print(
            f"ExpDataPack: 复制 Quantification-IS：{quant_is_src} -> {dest_path}")
        shutil.copy2(quant_is_src, dest_path)


def pack_scheme_gm(in_dir: str, exp_dir: str, grp_type: str | None, product_type: str):
    """
    大高靶方案二：肠菌300（GCMS+LCMS 或仅 LCMS）

    结构：
    exp-data/
      ├─ GCMS/...  （如果有）
      ├─ LCMS/...
      ├─ QC-corrplot-GCMS.jpg
      ├─ QC-corrplot-LCMS.jpg
      └─ Quantification-IS.xlsx
    """
    # 肠菌300-LC 只有 LCMS
    if "肠菌300-LC" in product_type or "GM300-LC" in product_type:
        grps = ["LCMS"]
    else:
        if not grp_type:
            grp_type = "GCMS|LCMS"
        grps = [g.strip() for g in str(grp_type).split("|") if g.strip()]

    if not grps:
        return

    for grp in grps:
        try:
            import re
            grp_in_dir = re.sub(
                grp_type if grp_type else "GCMS|LCMS", grp, in_dir)
        except Exception:
            grp_in_dir = in_dir

        grp_root = os.path.join(exp_dir, grp)
        pack_qc_spl_for_dir(grp_in_dir, grp_root)

        rt_dir = os.path.join(grp_root, "RT")
        ensure_dir(rt_dir)
        copy_rt_is_files(grp_in_dir, rt_dir)

    # QC-corrplot 图（放在 exp_dir 根目录，与 Quantification-IS.xlsx 同级）
    for grp in grps:
        try:
            import re
            grp_in_dir = re.sub(
                grp_type if grp_type else "GCMS|LCMS", grp, in_dir)
        except Exception:
            grp_in_dir = in_dir

        results_dir = os.path.join(grp_in_dir, "results")
        # 源文件已经带有 grp 后缀，直接查找完整文件名
        src_filename = f"QC-corrplot-{grp}.jpg"
        qc_corrplot_src = os.path.join(results_dir, "figures", src_filename)
        if os.path.isfile(qc_corrplot_src):
            dest_path = os.path.join(exp_dir, src_filename)
            print(
                f"ExpDataPack: 复制 QC-corrplot：{qc_corrplot_src} -> {dest_path}")
            shutil.copy2(qc_corrplot_src, dest_path)
        else:
            print(f"ExpDataPack: 未找到 QC-corrplot 图：{qc_corrplot_src}")

    # Quantification-IS.xlsx
    first_grp = grps[0]
    try:
        import re
        first_grp_dir = re.sub(
            grp_type if grp_type else "GCMS|LCMS", first_grp, in_dir)
    except Exception:
        first_grp_dir = in_dir

    quant_is_src = os.path.join(
        first_grp_dir, "results", "Quantification-IS.xlsx")
    if os.path.isfile(quant_is_src):
        dest_path = os.path.join(exp_dir, "Quantification-IS.xlsx")
        shutil.copy2(quant_is_src, dest_path)


def pack_scheme_exposome(in_dir: str, exp_dir: str):
    """
    大高靶方案三：暴露组

    结构：
    exp-data/
      ├─ TIC/QC/QC.png, SPL/SPL.png
      └─ QC/RT-Control-*.png, QC-corrplot.jpg, Quantification-IS.xlsx
    """
    # TIC 图
    tic_dir = os.path.join(exp_dir, "TIC")
    pack_tic_with_modes(in_dir, tic_dir, "QC", "QC")
    pack_tic_with_modes(in_dir, tic_dir, "SPL", "SPL")

    # QC 文件夹
    qc_dir = os.path.join(exp_dir, "QC")
    ensure_dir(qc_dir)

    # RT 图
    copy_rt_is_files(in_dir, qc_dir)

    # QC-corrplot
    results_dir = os.path.join(in_dir, "results")
    qc_corrplot_src = os.path.join(results_dir, "figures", "QC-corrplot.jpg")
    if os.path.isfile(qc_corrplot_src):
        dest_path = os.path.join(qc_dir, "QC-corrplot.jpg")
        shutil.copy2(qc_corrplot_src, dest_path)

    # Quantification-IS.xlsx
    quant_is_src = os.path.join(results_dir, "Quantification-IS.xlsx")
    if os.path.isfile(quant_is_src):
        dest_path = os.path.join(qc_dir, "Quantification-IS.xlsx")
        shutil.copy2(quant_is_src, dest_path)


def pack_scheme_lipids(in_dir: str, exp_dir: str):
    """
    定量脂质

    结构：
    exp-data/
      ├─ TIC/QC/QC-01__pos1.png, QC-01__neg.png
             SPL/SPL__pos1.png, SPL__neg.png
      └─ QC/RT-Control-*.png, QC-corrplot.jpg, Quantification-IS.xlsx
    """
    # TIC 图（带 pos/neg 模式）
    tic_dir = os.path.join(exp_dir, "TIC")
    pack_tic_with_modes(in_dir, tic_dir, "QC", "QC")
    pack_tic_with_modes(in_dir, tic_dir, "SPL", "SPL")

    # QC 文件夹
    qc_dir = os.path.join(exp_dir, "QC")
    ensure_dir(qc_dir)

    # RT 图
    copy_rt_is_files(in_dir, qc_dir)

    # QC-corrplot
    results_dir = os.path.join(in_dir, "results")
    qc_corrplot_src = os.path.join(results_dir, "figures", "QC-corrplot.jpg")
    if os.path.isfile(qc_corrplot_src):
        dest_path = os.path.join(qc_dir, "QC-corrplot.jpg")
        shutil.copy2(qc_corrplot_src, dest_path)

    # Quantification-IS.xlsx
    quant_is_src = os.path.join(results_dir, "Quantification-IS.xlsx")
    if os.path.isfile(quant_is_src):
        dest_path = os.path.join(qc_dir, "Quantification-IS.xlsx")
        shutil.copy2(quant_is_src, dest_path)


# ---------- 主打包逻辑 ----------

def pack_exp_data(in_dir: str):
    # 查找 headMessage 文件
    head_files = [f for f in os.listdir(in_dir) if f.startswith("headMessage")]
    if not head_files:
        print("ExpDataPack: 未找到 headMessage 文件，跳过打包。")
        return

    head_path = os.path.join(in_dir, head_files[0])

    # 读取首页 headMessage（索引在第一列），用 openpyxl 直接读
    try:
        wb = load_workbook(head_path, read_only=True, data_only=True)
        ws = wb.worksheets[0]

        # 找到 "项目类型" 行
        product_type = None
        for row in ws.iter_rows(min_row=1, max_row=50, values_only=True):
            if row[0] == "项目类型":
                product_type = row[1]
                break

        wb.close()

        if not product_type:
            print("ExpDataPack: headMessage 中未找到'项目类型'字段")
            return
    except Exception as e:
        print(f"ExpDataPack: 读取 headMessage 失败：{e}")
        return

    scheme = detect_scheme(product_type)
    if not scheme:
        print(f"ExpDataPack: 未识别的项目类型 '{product_type}'，不打包 exp-data。")
        return

    # 需要 GRP_TYPE 的方案：scheme3, scheme_aq_rp_hilic, scheme_gm
    grp_type = None
    if scheme in ["scheme3", "scheme_aq_rp_hilic", "scheme_gm"]:
        try:
            wb2 = load_workbook(head_path, read_only=True, data_only=True)
            # 找 "项目类型" sheet
            if "项目类型" in wb2.sheetnames:
                ws2 = wb2["项目类型"]
                # 第一行是表头，找 "合并方式[GRP_TYPE]" 列
                headers = [cell.value for cell in ws2[1]]
                if "合并方式[GRP_TYPE]" in headers:
                    col_idx = headers.index("合并方式[GRP_TYPE]") + 1
                    # 遍历找 product_type 行
                    for row in ws2.iter_rows(min_row=2, values_only=True):
                        if row[0] == product_type:
                            grp_type = row[col_idx - 1]
                            break
            wb2.close()
        except Exception as e:
            print(f"ExpDataPack: 读取项目类型配置失败：{e}")

    exp_dir = os.path.join(in_dir, "exp-data")
    ensure_dir(exp_dir)

    if scheme == "scheme1":
        pack_scheme1(in_dir, exp_dir)
    elif scheme == "scheme2":
        pack_scheme2(in_dir, exp_dir)
    elif scheme == "scheme3":
        pack_scheme3(in_dir, exp_dir, grp_type)
    elif scheme == "scheme_aq_rp_hilic":
        pack_scheme_aq_rp_hilic(in_dir, exp_dir, grp_type)
    elif scheme == "scheme_gm":
        pack_scheme_gm(in_dir, exp_dir, grp_type, product_type)
    elif scheme == "scheme_exposome":
        pack_scheme_exposome(in_dir, exp_dir)
    elif scheme == "scheme_lipids":
        pack_scheme_lipids(in_dir, exp_dir)

    # 生成 exp-data.zip（放在项目路径下）
    zip_path = os.path.join(in_dir, "exp-data.zip")
    if os.path.isfile(zip_path):
        os.remove(zip_path)

    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for root, dirs, files in os.walk(exp_dir):
            for name in files:
                abs_path = os.path.join(root, name)
                # 归档路径以项目根目录为基准，保证 zip 里是 exp-data/ 开头
                rel_path = os.path.relpath(abs_path, start=in_dir)
                zf.write(abs_path, arcname=rel_path)

    print(f"ExpDataPack: 已生成 {zip_path}")


def main():
    if len(sys.argv) < 2:
        print("用法：python ExpDataPack.py <项目路径>")
        sys.exit(1)

    in_dir = sys.argv[1]
    if not os.path.isdir(in_dir):
        print(f"错误：项目路径不存在：{in_dir}")
        sys.exit(1)

    pack_exp_data(in_dir)


if __name__ == "__main__":
    main()
